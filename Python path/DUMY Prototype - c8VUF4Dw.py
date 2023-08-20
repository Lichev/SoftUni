import openpyxl



dumy_file = openpyxl.load_workbook(path)

worksheet = dumy_file.active

stores_file = open("stores.txt", "w+", encoding="UTF-8")
storedtgroup_file = open("storeDTgroup.txt", "w+", encoding="UTF-8")
storexcodegroup_file = open("storeXgroup.txt", "w+", encoding="UTF-8")
lbatchstores_file = open("Lbatchstores.txt", "w+", encoding="UTF-8")

for i in range(1, worksheet.max_row):
    info_list = []
    for column in worksheet.iter_cols(1, 17):
        info_list.append(column[i].value)

    valid = info_list[0] is None
    DUMY_ID = "53MO" + info_list[1]
    AC_NSHOPID = info_list[2]
    AC_SHOPDESCRIPTION = info_list[3]
    AC_SHOPSTATUS = info_list[4]
    AC_DEFAULTXCODEGR = info_list[5]
    MM_XCODEGROUP = AC_DEFAULTXCODEGR + "_M"
    AC_COUNTRYID = info_list[6]
    AC_RETAILER = info_list[7]
    AC_LANGUAGEID = info_list[8]
    AC_SHOPTYPE = info_list[9]
    AC_AREA = info_list[10]
    NC_SURFACE = info_list[11]
    NC_ACTIVEFLAG = info_list[12]
    NC_DUPITEMS_FLAG = info_list[13]
    AC_CHANNELID = info_list[14]
    NC_DUMMY_FLAG = info_list[15]
    AC_STORE_CHAR1 = info_list[16]

    if valid:

        # stores table starts here
        stores_list = []

        stores_list.append(DUMY_ID)
        stores_list.append('WEK_DUMMY')
        stores_list.append('REQUIRED')
        stores_list.append(AC_DEFAULTXCODEGR)
        stores_list.append("RS")
        stores_list.append('DUMMY')
        stores_list.append('SR')
        stores_list.append('AUDIT')
        stores_list.append(AC_AREA)
        stores_list.append(NC_SURFACE)
        stores_list.append('100')
        stores_list.append('1')
        stores_list.append('1')
        stores_list.append('1')
        stores_list.append('0')
        stores_list.append('MONT')
        stores_list.append('0')
        stores_list.append(AC_STORE_CHAR1)
        stores_list.append('MNTL')

        stores_print = '\t'.join(str(e) for e in stores_list)
        stores_file.write(stores_print)
        stores_file.write("\n")


        # storeDTgroup starts here
        storedtgroup_list = []

        storedtgroup_list.append(DUMY_ID)
        storedtgroup_list.extend(('AUDIT_DTYPE', '0', '1'))

        storedtgroup_print = '\t'.join(str(dtgroup) for dtgroup in storedtgroup_list)
        storedtgroup_file.write(storedtgroup_print)
        storedtgroup_file.write("\n")

        #storeXcodegroup start here
        first_level = [DUMY_ID, DUMY_ID, '1', '0', '9999']
        second_level = [DUMY_ID, MM_XCODEGROUP, '2', '0', '9999']
        third_level = [DUMY_ID, 'RSXX', '3', '0', '9999']
        fourth_level = [DUMY_ID, 'NAN_KEY', '4', '0', '9999']

        first_print = '\t'.join(str(one) for one in first_level)
        second_print = '\t'.join(str(two) for two in second_level)
        third_print = '\t'.join(str(three) for three in third_level)
        fourth_print = '\t'.join(str(four) for four in fourth_level)

        storexcodegroup_file.write(first_print)
        storexcodegroup_file.write("\n")
        storexcodegroup_file.write(second_print)
        storexcodegroup_file.write("\n")
        storexcodegroup_file.write(third_print)
        storexcodegroup_file.write("\n")
        storexcodegroup_file.write(fourth_print)
        storexcodegroup_file.write("\n")


        #lbatchstores starts here:
        lbatchstores_list = ['SISTERSTORES_WEEKLY-MONTHLY_2', DUMY_ID, AC_NSHOPID, '0', '9999']

        lbatchstores_print = '\t'.join(str(lbatch) for lbatch in lbatchstores_list)
        lbatchstores_file.write(lbatchstores_print)
        lbatchstores_file.write('\n')

    else:
        continue

stores_file.close()
storedtgroup_file.close()
storexcodegroup_file.close()
lbatchstores_file.close()
