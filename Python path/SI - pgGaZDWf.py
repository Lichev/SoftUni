import openpyxl

print('Welcome to Slovenia new store creation tool:')
language = input("Path language: ")
name = input("Please enter your login name: ")

if language == "BG" or language == 'bg':
    
else:
    


def type_of_store(area):
    if area == 'DMM' or area == 'ODR' or area == 'TDR':
        return 'DRUG CHAIN'
    elif area == 'SSU' or area == 'MSU' or area == 'MHY' or area == 'MLA' or area == 'SHY' or area == 'MME' or \
            area == 'OLA' or area == "OME" or area == 'OSM' or area == 'OSU' or area == 'DIS' or area == 'SLA' \
            or area == 'OHY' or area == 'THY' or area == 'TSU' or area == 'TLA':
        return 'GROCERY'
    elif area == 'OKS':
        return 'KIOSKS'
    elif area == 'OMV' or area == 'PET' or area == 'MOL' or area == 'GAS':
        return 'PETROL STATIONS'


def print_store(numbers):
    wb_obj = openpyxl.load_workbook(path)

    worksheet = wb_obj.active

    row_min = worksheet.max_row - numbers

    big_list = []

    stores_file = open("stores.txt", "w+")
    storedtgroup_file = open("storeDTgroup.txt", "w+")
    storexcodegroup_file = open("storeXgroup.txt", "w+")
    lbatchstores_file = open("Lbatchstores.txt", "w+")

    # taking info from excel
    for i in range(row_min, worksheet.max_row):
        temp_list = []
        for col in worksheet.iter_cols(1, 12):
            temp_list.append(col[i].value)

        # taking data by postion
        batch = temp_list[1]
        type_of_data = temp_list[2]
        shop_id = temp_list[4]
        ac_nshopid = "5100" + shop_id
        xcodegr = temp_list[5]
        shopdescr = temp_list[7]
        retailer = temp_list[8]
        area = temp_list[9]
        surface = temp_list[10]
        region = temp_list[11]
        ac_shopstatus = 'REQUIRED'
        ac_countryid = 'SI'
        ac_languageid = 'EN'
        ac_shoptype = type_of_store(area)
        nc_acv = 100
        nc_xf = 1
        nc_activeflag = 1
        nc_dupitems_flag = 1
        nc_eanxcode_flag = 0
        ac_channelid = 'SCAN'
        nc_dumy_flag = 0
        ac_store_char2 = 'WKLY'

        # create stores list
        final_list = []

        final_list.append(ac_nshopid)
        final_list.append(shopdescr)
        final_list.append(ac_shopstatus)
        final_list.append(xcodegr)
        final_list.append(ac_countryid)
        final_list.append(retailer)
        final_list.append(ac_languageid)
        final_list.append(ac_shoptype)
        final_list.append(area)
        final_list.append(surface)
        final_list.append(nc_acv)
        final_list.append(nc_xf)
        final_list.append(nc_activeflag)
        final_list.append(nc_dupitems_flag)
        final_list.append(nc_eanxcode_flag)
        final_list.append(ac_channelid)
        final_list.append(nc_dumy_flag)
        final_list.append(region)

        # print stores list into notepad
        printer_stores = '\t'.join(str(e) for e in final_list)
        stores_file.write(printer_stores)
        stores_file.write("\n")

        # create StoreDT group list
        store_dt_group = []

        store_dt_group.append(ac_nshopid)
        store_dt_group.extend(('VOLUMETRIC', '0', '1'))
        printer_storedtgroup = '\t'.join(str(a) for a in store_dt_group)
        storedtgroup_file.write(printer_storedtgroup)
        storedtgroup_file.write("\n")

        # create StoreXcode groups list
        level_one = []
        level_two = []
        level_three = []
        level_four = []

        # level one
        level_one.append(ac_nshopid)
        level_one.append(ac_nshopid)
        level_one.extend(('1', '0', '9999'))

        # level two
        level_two.append(ac_nshopid)
        level_two.append(xcodegr)
        level_two.extend(('2', '0', '9999'))

        # level three
        level_three.append(ac_nshopid)
        level_three.append('SIXX')
        level_three.extend(('3', '0', '9999'))

        # level four
        level_four.append(ac_nshopid)
        level_four.append('EAN')
        level_four.extend(('4', '0', '9999'))

        printer_storeXcode_one = '\t'.join(str(one) for one in level_one)
        printer_storeXcode_two = '\t'.join(str(two) for two in level_two)
        printer_storeXcode_three = '\t'.join(str(three) for three in level_three)
        printer_storeXcode_four = '\t'.join(str(four) for four in level_four)
        storexcodegroup_file.write(printer_storeXcode_one)
        storexcodegroup_file.write("\n")
        storexcodegroup_file.write(printer_storeXcode_two)
        storexcodegroup_file.write("\n")
        storexcodegroup_file.write(printer_storeXcode_three)
        storexcodegroup_file.write("\n")
        storexcodegroup_file.write(printer_storeXcode_four)
        storexcodegroup_file.write("\n")

        # create lbatchstores

        lbatchstores = []

        lbatchstores.append(batch)
        lbatchstores.append(shop_id)
        lbatchstores.append(ac_nshopid)
        lbatchstores.extend(('0', '1'))

        printer_lbatchstores = '\t'.join(str(l) for l in lbatchstores)
        lbatchstores_file.write(printer_lbatchstores)
        lbatchstores_file.write("\n")

    stores_file.close()
    storedtgroup_file.close()
    storexcodegroup_file.close()
    lbatchstores_file.close()


lines = int(input("Number of lines: "))
print_store(lines)
