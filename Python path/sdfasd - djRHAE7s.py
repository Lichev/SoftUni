import openpyxl

print('Welcome to SERBIA new store creation tool:')




def type_of_store(shop_type):
    if shop_type == 'SUP' or shop_type == 'HYP' or shop_type == 'FONL':
        return 'SUPER_HYPER'
    elif shop_type == 'MCC':
        return 'CASH & CARRY'
    elif shop_type == 'DIS':
        return 'DISCOUNTER'
    elif shop_type == 'LGR' or shop_type == 'MGR' or shop_type == 'SGR' or shop_type == 'OAL':
        return 'GROCERY'
    elif shop_type == 'MXK' or shop_type == 'PAV' or shop_type == '1XK' or shop_type == '1AV':
        return 'KIOSK'
    elif shop_type == 'PTS':
        return 'PETROL STATION'
    elif shop_type == 'DRG' or shop_type == 'BAB' or shop_type == 'PFM' or shop_type == 'ONL':
        return 'SPECIALIZED DRUG STORE'


def print_store(numbers):
    wb_obj = openpyxl.load_workbook(path)

    worksheet = wb_obj.active

    row_min = worksheet.max_row - numbers

    big_list = []

    stores_file = open("stores.txt", "w+")
    storedtgroup_file = open("storeDTgroup.txt", "w+")
    storexcodegroup_file = open("storeXgroup.txt", "w+")
    lbatchstores_file = open("Lbatchstores.txt", "w+")

    # taking info from excel
    for i in range(row_min, worksheet.max_row):
        temp_list = []
        for col in worksheet.iter_cols(1, 12):
            temp_list.append(col[i].value)

        # taking data by postion
        batch = temp_list[1]
        type_of_data = temp_list[2]
        shop_id = temp_list[4]
        xcodegr = temp_list[5]
        shopdescr = temp_list[7]
        retailer = temp_list[8]
        area = temp_list[9]
        surface = temp_list[10]
        region = temp_list[11]
        ac_shopstatus = 'REQUIRED'
        ac_countryid = 'RS'
        ac_languageid = 'SR'
        ac_shoptype = type_of_store(area)
        nc_acv = 100
        nc_xf = 1
        nc_activeflag = 1
        nc_dupitems_flag = 1
        nc_eanxcode_flag = 0
        ac_channelid = 'SCAN'
        nc_dumy_flag = 0
        ac_store_char2 = 'WKLY'

        if type_of_data == 'SCANNING':
            # create stores list
            final_list = []

            final_list.append('5300' + shop_id)
            final_list.append(shopdescr)
            final_list.append(ac_shopstatus)
            final_list.append(xcodegr)
            final_list.append(ac_countryid)
            final_list.append(retailer)
            final_list.append(ac_languageid)
            final_list.append(ac_shoptype)
            final_list.append(area)
            final_list.append(surface)
            final_list.append(nc_acv)
            final_list.append(nc_xf)
            final_list.append(nc_activeflag)
            final_list.append(nc_dupitems_flag)
            final_list.append(nc_eanxcode_flag)
            final_list.append(ac_channelid)
            final_list.append(nc_dumy_flag)
            final_list.append(region)
            final_list.append(ac_store_char2)

            # print stores list into notepad
            printer_stores = '\t'.join(str(e) for e in final_list)
            stores_file.write(printer_stores)
            stores_file.write("\n")

            # create StoreDT group list
            store_dt_group = []

            store_dt_group.append('5300' + shop_id)
            store_dt_group.extend(('VOLUMETRIC', '0', '1'))
            printer_storedtgroup = '\t'.join(str(a) for a in store_dt_group)
            storedtgroup_file.write(printer_storedtgroup)
            storedtgroup_file.write("\n")

            # create StoreXcode groups list
            level_one = []
            level_two = []
            level_three = []
            level_four = []

            # level one
            level_one.append('5300' + shop_id)
            level_one.append('5300' + shop_id)
            level_one.extend(('1', '0', '9999'))

            # level two
            level_two.append('5300' + shop_id)
            level_two.append(xcodegr)
            level_two.extend(('2', '0', '9999'))

            # level three
            level_three.append('5300' + shop_id)
            level_three.append('RSXX')
            level_three.extend(('3', '0', '9999'))

            # level four
            level_four.append('5300' + shop_id)
            level_four.append('EAN')
            level_four.extend(('4', '0', '9999'))

            printer_storeXcode_one = '\t'.join(str(one) for one in level_one)
            printer_storeXcode_two = '\t'.join(str(two) for two in level_two)
            printer_storeXcode_three = '\t'.join(str(three) for three in level_three)
            printer_storeXcode_four = '\t'.join(str(four) for four in level_four)
            storexcodegroup_file.write(printer_storeXcode_one)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(printer_storeXcode_two)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(printer_storeXcode_three)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(printer_storeXcode_four)
            storexcodegroup_file.write("\n")

            # create lbatchstores

            lbatchstores = []

            lbatchstores.append(batch)
            lbatchstores.append(shop_id)
            lbatchstores.append('5300' + shop_id)
            lbatchstores.extend(('0', '1'))

            printer_lbatchstores = '\t'.join(str(l) for l in lbatchstores)
            lbatchstores_file.write(printer_lbatchstores)
            lbatchstores_file.write("\n")



        elif type_of_data == 'A2S':
            scan_list = []
            monthly_list = []

            scan_list.append('53AS' + shop_id)
            scan_list.append(shopdescr)
            scan_list.append(ac_shopstatus)
            scan_list.append(xcodegr)
            scan_list.append(ac_countryid)
            scan_list.append(retailer)
            scan_list.append(ac_languageid)
            scan_list.append(ac_shoptype)
            scan_list.append(area)
            scan_list.append(surface)
            scan_list.append(nc_acv)
            scan_list.append(nc_xf)
            scan_list.append(nc_activeflag)
            scan_list.append(nc_dupitems_flag)
            scan_list.append(nc_eanxcode_flag)
            scan_list.append('DUMY')
            scan_list.append(1)
            scan_list.append(region)
            scan_list.append(ac_store_char2)
            scan_list.append('A2S DUMMY')

            monthly_list.append('53MO' + shop_id)
            monthly_list.append(shopdescr)
            monthly_list.append(ac_shopstatus)
            monthly_list.append(xcodegr)
            monthly_list.append(ac_countryid)
            monthly_list.append(retailer)
            monthly_list.append(ac_languageid)
            monthly_list.append("AUDIT")
            monthly_list.append(area)
            monthly_list.append(surface)
            monthly_list.append(nc_acv)
            monthly_list.append(nc_xf)
            monthly_list.append(nc_activeflag)
            monthly_list.append(nc_dupitems_flag)
            monthly_list.append(nc_eanxcode_flag)
            monthly_list.append('MONТ')
            monthly_list.append(nc_dumy_flag)
            monthly_list.append(region)
            monthly_list.append('')
            monthly_list.append('A2S')

            # print stores list into notepad
            printer_stores = '\t'.join(str(e) for e in scan_list)
            print_mont = '\t'.join(str(m) for m in monthly_list)
            stores_file.write(printer_stores)
            stores_file.write("\n")
            stores_file.write(print_mont)
            stores_file.write("\n")


            #Create storedtgroup
            store_dt_group = []
            month_storedtgroup = []

            store_dt_group.append('53AS' + shop_id)
            store_dt_group.extend(('VOLUMETRIC', '0', '1'))
            month_storedtgroup.append(('53MO' + shop_id))
            month_storedtgroup.extend(('AUDIT_DTYPE', '0', '1'))

            printer_storedtgroup = '\t'.join(str(a) for a in store_dt_group)
            print_monthdtgroup = '\t'.join(str(m) for m in month_storedtgroup)
            storedtgroup_file.write(printer_storedtgroup)
            storedtgroup_file.write("\n")
            storedtgroup_file.write(print_monthdtgroup)
            storedtgroup_file.write("\n")

            # create StoreXcode groups list
            level_one = []
            level_two = []
            level_three = []
            level_four = []
            mm_one = []
            mm_two = []
            mm_three = []
            mm_four = []

            # level one weekly code
            level_one.append('53AS' + shop_id)
            level_one.append('53AS' + shop_id)
            level_one.extend(('1', '0', '9999'))

            # level two weekly code
            level_two.append('53AS' + shop_id)
            level_two.append(xcodegr)
            level_two.extend(('2', '0', '9999'))

            # level three weekly code
            level_three.append('53AS' + shop_id)
            level_three.append('RSXX')
            level_three.extend(('3', '0', '9999'))

            # level four weekly code
            level_four.append('53AS' + shop_id)
            level_four.append('EAN')
            level_four.extend(('4', '0', '9999'))

            # level one monthly code
            mm_one.append('53MO' + shop_id)
            mm_one.append('53MO' + shop_id)
            mm_one.extend(('1', '0', '9999'))

            # level two monthly code
            mm_two.append('53MO' + shop_id)
            mm_two.append(xcodegr + '_M')
            mm_two.extend(('2', '0', '9999'))

            # level three monthly code
            mm_three.append('53MO' + shop_id)
            mm_three.append('RSXX')
            mm_three.extend(('3', '0', '9999'))

            # level three monthly code
            mm_four.append('53MO' + shop_id)
            mm_four.append('NAN_KEY')
            mm_four.extend(('4', '0', '9999'))

            print_wk_one = '\t'.join(str(one) for one in level_one)
            print_wk_two = '\t'.join(str(two) for two in level_two)
            print_wk_three = '\t'.join(str(three) for three in level_three)
            print_wk_four = '\t'.join(str(four) for four in level_four)
            print_mm_one = '\t'.join(str(mone) for mone in mm_one)
            print_mm_two = '\t'.join(str(mtwo) for mtwo in mm_two)
            print_mm_three = '\t'.join(str(mthree) for mthree in mm_three)
            print_mm_four = '\t'.join(str(mfour) for mfour in mm_four)

            storexcodegroup_file.write(print_wk_one)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(print_wk_two)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(print_wk_three)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(print_wk_four)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(print_mm_one)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(print_mm_two)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(print_mm_three)
            storexcodegroup_file.write("\n")
            storexcodegroup_file.write(print_mm_four)
            storexcodegroup_file.write("\n")


            # create lbatchstores_wk

            lbatchstores_wk = []
            lbatchstores_mm = []

            lbatchstores_wk.append(batch)
            lbatchstores_wk.append(shop_id)
            lbatchstores_wk.append('53AS' + shop_id)
            lbatchstores_wk.extend(('0', '1'))

            lbatchstores_mm.append('SISTERSTORES_WEEKLY-MONTHLY')
            lbatchstores_mm.append('53MO' + shop_id)
            lbatchstores_mm.append('53AS' + shop_id)
            lbatchstores_mm.extend(('0', '1'))

            printer_lbatchstores = '\t'.join(str(l) for l in lbatchstores_wk)
            printer_lbatchstores_mm = '\t'.join(str(mm) for mm in lbatchstores_mm)
            lbatchstores_file.write(printer_lbatchstores)
            lbatchstores_file.write("\n")
            lbatchstores_file.write(printer_lbatchstores_mm)
            lbatchstores_file.write("\n")


    stores_file.close()
    storedtgroup_file.close()
    storexcodegroup_file.close()
    lbatchstores_file.close()


lines = int(input("Number of lines: "))
print_store(lines)
