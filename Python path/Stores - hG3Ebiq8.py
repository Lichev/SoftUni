import numbers

import openpyxl
import sys

path = "D:\\NewStores.xlsx"


def type_of_store(shop_type):
    if shop_type == 'SUP' or shop_type == 'HYP' or shop_type == 'FONL':
        return 'SUPER_HYPER'
    elif shop_type == 'MCC':
        return 'CASH & CARRY'
    elif shop_type == 'DIS':
        return 'DISCOUNTER'
    elif shop_type == 'LGR' or shop_type == 'MGR' or shop_type == 'SGR' or shop_type == 'OAL':
        return 'GROCERY'
    elif shop_type == 'MXK' or shop_type == 'PAV' or shop_type == '1XK' or shop_type == '1AV':
        return 'KIOSK'
    elif shop_type == 'PTS':
        return 'PETROL STATION'
    elif shop_type == 'DRG' or shop_type == 'BAB' or shop_type == 'PFM' or shop_type == 'ONL':
        return 'SPECIALIZED DRUG STORE'


def print_store(numbers):
    wb_obj = openpyxl.load_workbook(path)

    worksheet = wb_obj.active

    row_min = worksheet.max_row - numbers

    big_list = []

    f = open("stores.txt", "w+")

    for i in range(row_min, worksheet.max_row):
        temp_list = []
        for col in worksheet.iter_cols(1, 12):
            temp_list.append(col[i].value)

        batch = temp_list[1]
        type_of_data = temp_list[2]
        shop_id = temp_list[4]
        xcodegr = temp_list[5]
        shopdescr = temp_list[7]
        retailer = temp_list[8]
        area = temp_list[9]
        surface = temp_list[10]
        region = temp_list[11]

        final_list = []

        final_list.append('5300' + shop_id)
        final_list.append('REQUIRED')
        final_list.append(xcodegr)
        final_list.extend(('RS', 'SR', '100', '1', '1', '0', "SCAN", '0'))
        final_list.append(shopdescr)
        final_list.append(retailer)
        final_list.append(area)
        final_list.append(type_of_store(area))
        final_list.append(surface)
        final_list.append('1')
        final_list.append('WKLY')

        print(*final_list, sep="    ")


lines = int(input())
print_store(lines)

